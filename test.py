# ==================================================
# define
# ==================================================
FILE_NAME               = "locale.xlsx"
OUTPUT_JSON_FILE_NAME   = "locale.json"
VERSION                 = "0.1"

APP_SHEET = "シート1"
KEY_NAME = "ID"
LOCALE_LIST = [
	"テキスト_0", 
	"テキスト_1"
]

LOG_HEAD = "    "
# ==================================================
# import
# ==================================================
import sys
import openpyxl
import json
import os
from tqdm import tqdm
import time

# ==================================================
# api
# ==================================================
def log( arg ):
    print(LOG_HEAD + str(arg))

def isExistSheet( wb, sheet_name ):
	result = False
	for ws in wb.worksheets:
	    if ws.title == sheet_name:
	        result = True
	        break
	return result

def getKeyCellCoordinate( ws, key_name ):
	result_row = -1
	result_col = -1
	for r in ws.iter_rows(max_row=100, max_col=0):
		for c in r:
			if c.value == KEY_NAME:
				cell_nloc=c.coordinate
				result_row=openpyxl.utils.cell.coordinate_from_string(cell_nloc)[1] + 1
				result_col=openpyxl.utils.column_index_from_string(cell_nloc[0]) - 1
	return result_row, result_col

# ==================================================
# main
# ==================================================
print()
log("Start convert")
print()

#ファイル存在確認
if os.path.exists(FILE_NAME) == False:
	log(FILE_NAME + 'が存在しません')
	sys.exit()

log("File\t: " + FILE_NAME)

#Excelファイルの読み込み
wb = openpyxl.load_workbook(FILE_NAME)

#シート存在確認
if isExistSheet(wb, APP_SHEET) == False:
    log(APP_SHEET + 'が存在しません')
    sys.exit()

log("Sheet\t: " + APP_SHEET)

#シートの取得
ws = wb[APP_SHEET]

#Keyの列番号取得
key_row_index, key_colum_index = getKeyCellCoordinate(ws, KEY_NAME)
if key_colum_index == -1:
    log("Key: " + KEY_NAME + 'が存在しません')
    sys.exit()

#各言語の列番号辞書
locale_colum_index_dic = {}

#各言語の列番号取得
for locale_str in LOCALE_LIST:
	for r in ws.iter_rows(max_row=100, max_col=0):
		for c in r:
			if c.value == locale_str:
				cell_nloc=c.coordinate
				cell_col_m=openpyxl.utils.column_index_from_string(cell_nloc[0])
				locale_colum_index_dic[locale_str] = cell_col_m-1

#各言語テキスト辞書
locale_message_dic = {}
for locale_str in LOCALE_LIST:
	locale_message_dic[locale_str] = {}

#各言語テキストの取得
for row in tqdm(ws.iter_rows(min_row = key_row_index, min_col = 0, max_col = 100), leave=False):
	values = [cell.value for cell in row]
	if values[0] is None: break
	for locale_str in LOCALE_LIST:
		locale_message_dic[locale_str][values[key_colum_index]] = values[locale_colum_index_dic[locale_str]]
	time.sleep(0.1)

#出力するjsonファイルデータ
data_dict = {
    "version": VERSION,
    "data": locale_message_dic
}

#jsonファイル出力
with open(OUTPUT_JSON_FILE_NAME, mode = 'w', encoding = 'utf-8') as f:
    f.write(json.dumps(data_dict, ensure_ascii = False, indent = 4))

log("OUTPUT\t: " + os.path.abspath(OUTPUT_JSON_FILE_NAME))

print()
log("Finished")


